#!/usr/bin/env python
# coding: utf-8

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import warnings
import seaborn as sns
import time
import os
from sklearn.ensemble import RandomForestRegressor
from openpyxl import load_workbook
warnings.filterwarnings("ignore")


class calculate_IV:
    def __init__(self):
        import pandas as pd
        import numpy as np
        pd.set_option("display.max_columns",None)
        
    def _createBin(self):
        try:
            self.max_bin = int(input("Enter the Number of Bin (not more than 30): "))
            if self.max_bin <= 30 and self.max_bin > 0:
                return self.max_bin
            else:
                self.max_bin = 3
                return self.max_bin
        
        except:
            print("Max_bin creation failure:")
            
    def __main__(self,data,target):
        self.bin = self._createBin()
        print(self.bin)
        from pandas import Series
        x = list(data.dtypes.index)
        count = -1
        for i in x:
            if np.issubdtype(data[i], np.number) and len(Series.unique(data[i])) > 2:
                print("variable type: Numeric == ",i)
                self.con_var = self._monoBin(target,data[i],self.bin,i)
                #print (self.con_var)
                
            else:
                print("Variable type: categorical == ",i)
                self.cat_var = self._charBin(target,data[i],i)
                #print(self.cat_var)
            
            
    def _monoBin(self,y,X,bin_size,var_name):
        #self.bin = self._createBin()
        import scipy.stats.stats as stats
        data_temp = pd.DataFrame({"X":X, "Y":y})
        missing_data = data_temp[['X','Y']][data_temp['X'].isnull()]
        non_missing_data = data_temp[['X','Y']][data_temp['X'].notnull()]
        
        start_row = 0
        self.sheet_name = var_name #imp
        while np.abs(bin_size) >= 3:
            d1 = pd.DataFrame({"X":non_missing_data.X,
                               "Y" : non_missing_data.Y,
                               "Bucket" : pd.cut(non_missing_data.X, bin_size)
                              })
            d2 = d1.groupby('Bucket', as_index = True)
            d3 = pd.DataFrame({}, index = [])
            d3['MIN_VALUE'] = d2.min().X
            d3['MAX_VALUE'] = d2.max().X
            d3['MEAN_VALUE'] = d2.mean().X
            d3['COUNT'] = d2.count().Y
            d3['EVENT'] = d2.sum().Y
            d3['NON_EVENT'] = d2.count().Y - d2.sum().Y
            d3 = d3.reset_index(drop = True)
            
            if len(missing_data.index) > 0:
                d4 = pd.DataFrame({'MIN_VALUE': np.nan}, index =[0])
                d4['MAX_VALUE'] = np.nan
                d4['MEAN_VALUE'] = np.nan
                d4['COUNT'] = missing_data.count().Y
                d4['EVENT'] = missing_data.sum().Y
                d4['NON_EVENT'] = missing_data.count().Y - missing_data.sum().Y
                d3 = d3.append(d4, ignore_index = True)
            
            d3['EVENT_RATE'] = d3['EVENT']/d3['COUNT']
            d3['NON_EVENT_RATE'] = d3['NON_EVENT']/d3['COUNT']
            d3["DIST_EVENT"] = d3.EVENT/d3.sum().EVENT
            d3["DIST_NON_EVENT"] = d3.NON_EVENT/d3.sum().NON_EVENT
            d3["WOE"] = np.log(d3.DIST_EVENT/d3.DIST_NON_EVENT)
            d3["IV"] = (d3.DIST_EVENT-d3.DIST_NON_EVENT)*np.log(d3.DIST_EVENT/d3.DIST_NON_EVENT)
            d3 = d3.replace([np.inf,-np.inf],0)
            d3 = d3.reset_index(drop = True)
            
            printing_statement = "IV value for bin size: " + str(bin_size)+ "  is: "+ str(np.round(d3.sum().IV,3))
            print(printing_statement)
            
            '''
            Now writing the output to an excel sheet
            '''          
            if not os.path.isfile("E:\IV output\output.xlsx"):          #update the local directore name and file name
                d3.to_excel("E:\IV output\output.xlsx",                 #update the local directore name and file name
                            sheet_name = self.sheet_name,
                            startrow=start_row if start_row is not None else 0,
                            index = False)
            else:
                writer = pd.ExcelWriter("E:\IV output\output.xlsx", engine='openpyxl', mode = 'a')
                writer.book = load_workbook("E:\IV output\output.xlsx")  #update the local directore name and file name
                writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
                d3.to_excel(writer, sheet_name = self.sheet_name, startrow = start_row,index = False)
                writer.save()
                start_row = start_row + bin_size + 3
            bin_size = bin_size - 1
    
    
    def _charBin(self,Y,X,var_name):
        self.sheet_name = var_name
        start_row = 0
        df_char = pd.DataFrame({"X":X,"Y":Y})
        char_miss_data = df_char[['X','Y']][df_char.X.isnull()]
        char_non_miss_data = df_char[["X","Y"]][df_char.X.notnull()]
        df2 = char_non_miss_data.groupby('X', as_index = True)
        
        d3 = pd.DataFrame({},index=[])
        d3['COUNT'] = df2.count().Y
        d3['VAR_TAG'] = var_name
        d3['EVENT'] = df2.sum().Y
        d3['NON_EVENT'] = df2.count().Y-df2.sum().Y
        d3['BUCKET'] = df2.groups.keys()
        
        if len(char_miss_data.index) > 0:
            d4 = pd.DataFrame({'BUCKET' : 'Missing'}, index = [0])
            d4['COUNT'] = char_miss_data.count().Y
            d4['EVENT'] = char_miss_data.sum().Y
            d4['NON_EVENT'] = char_miss_data.count().Y - char_miss_data.sum().Y
            d3 = d3.append(d4, ignore_index = True)
            
        d3['EVENT_RATE'] = d3.EVENT/d3.COUNT
        d3['NON_EVENT_RATE'] = d3.NON_EVENT/d3.COUNT
        d3['DIST_EVENT'] = d3.EVENT/d3.sum().EVENT
        d3['DIST_NON_EVENT'] = d3.NON_EVENT/d3.sum().NON_EVENT
        d3['WOE'] = np.log(d3.DIST_EVENT/d3.DIST_NON_EVENT)
        d3['IV'] = (d3.DIST_EVENT - d3.DIST_NON_EVENT)*np.log(d3.DIST_EVENT/d3.DIST_NON_EVENT)
        
        d3 = d3[['VAR_TAG','BUCKET','COUNT','EVENT','NON_EVENT',
                 'EVENT_RATE','NON_EVENT_RATE','DIST_EVENT','DIST_NON_EVENT',
                 'WOE','IV']]
        d3 = d3.replace([np.inf,-np.inf],0)
        d3 = d3.reset_index(drop = True)
        printing_statement = "IV Value for: " + self.sheet_name + " is: " + str(np.round(d3.sum().IV,3))
        print(printing_statement)
        '''
        Now writing the output to an excel sheet
        '''          
        if not os.path.isfile("E:\IV output\output.xlsx"):              #update the local directore name and file name
            d3.to_excel("E:\IV output\output.xlsx",                     #update the local directore name and file name
                        sheet_name = self.sheet_name,
                        startrow=start_row if start_row is not None else 0,
                        index = False)
        else:
            writer = pd.ExcelWriter("E:\IV output\output.xlsx", engine='openpyxl', mode = 'a')  #update the local directore name and file name
            writer.book = load_workbook("E:\IV output\output.xlsx")
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
            d3.to_excel(writer, sheet_name = self.sheet_name, startrow = start_row,index = False)
            writer.save()
'''
End of class IV_WOE
'''
            
            
#Example-1 for data selection prior to fitting to the main model
#Importing Titanic data data
df = pd.read_csv("E:\\ROB\\TitanicDataAnalysis\\train.csv")
X = df.drop("Survived",axis = 1)
X = X.drop("PassengerId",axis = 1)
y = df['Survived']

X_train, X_test, y_train, y_test = train_test_split(X,y, train_size = 0.80, random_state = 43)

d = calculate_IV()
d.__main__(X,y)

